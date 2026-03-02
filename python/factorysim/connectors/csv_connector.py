"""
CSV Connector - Import/export data from CSV and Excel files.
"""

from typing import Dict, Any, List, Optional
from pathlib import Path
import csv
import json


class CSVConnector:
    """
    Connector for importing and exporting CSV and Excel files.
    """

    def __init__(self):
        self.supported_formats = [".csv", ".xlsx", ".xls"]

    def import_file(self, file_path: str, options: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        Import data from a CSV file.

        Args:
            file_path: Path to the CSV file
            options: Import options (column_mapping, skip_rows, etc.)

        Returns:
            Dictionary with import results
        """
        options = options or {}

        try:
            path = Path(file_path)

            if not path.exists():
                return {
                    "success": False,
                    "error": f"File not found: {file_path}",
                    "rows_imported": 0,
                    "errors": [{"row": 0, "column": "", "message": "File not found"}],
                    "warnings": [],
                }

            if path.suffix.lower() not in self.supported_formats:
                return {
                    "success": False,
                    "error": f"Unsupported file format: {path.suffix}",
                    "rows_imported": 0,
                    "errors": [{"row": 0, "column": "", "message": f"Unsupported format: {path.suffix}"}],
                    "warnings": [],
                }

            # Read CSV file
            rows = []
            errors = []
            warnings = []

            skip_rows = options.get("skip_rows", 0)
            column_mapping = options.get("column_mapping", {})

            with open(file_path, 'r', newline='', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []

                # Apply column mapping
                if column_mapping:
                    mapped_headers = [column_mapping.get(h, h) for h in headers]
                else:
                    mapped_headers = headers

                for i, row in enumerate(reader):
                    if i < skip_rows:
                        continue

                    try:
                        # Map columns
                        mapped_row = {}
                        for original, mapped in zip(headers, mapped_headers):
                            value = row.get(original, "")
                            # Try to convert to number
                            try:
                                if "." in str(value):
                                    mapped_row[mapped] = float(value)
                                else:
                                    mapped_row[mapped] = int(value)
                            except (ValueError, TypeError):
                                mapped_row[mapped] = value

                        rows.append(mapped_row)

                    except Exception as e:
                        errors.append({
                            "row": i + 1,
                            "column": "",
                            "message": str(e),
                        })

            # Calculate quality score
            quality_score = 1.0 - (len(errors) / max(len(rows), 1))

            return {
                "success": len(errors) == 0,
                "rows_imported": len(rows),
                "data": rows,
                "headers": mapped_headers,
                "quality_score": quality_score,
                "errors": errors,
                "warnings": warnings,
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "rows_imported": 0,
                "errors": [{"row": 0, "column": "", "message": str(e)}],
                "warnings": [],
            }

    def import_excel(self, file_path: str, options: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        Import data from an Excel file.

        Args:
            file_path: Path to the Excel file
            options: Import options (sheet_name, column_mapping, skip_rows, etc.)

        Returns:
            Dictionary with import results
        """
        options = options or {}

        try:
            import openpyxl

            path = Path(file_path)
            if not path.exists():
                return {
                    "success": False,
                    "error": f"File not found: {file_path}",
                    "rows_imported": 0,
                    "errors": [{"row": 0, "column": "", "message": "File not found"}],
                    "warnings": [],
                }

            # Load workbook
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            # Get sheet
            sheet_name = options.get("sheet_name")
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return {
                        "success": False,
                        "error": f"Sheet not found: {sheet_name}",
                        "rows_imported": 0,
                        "errors": [{"row": 0, "column": "", "message": f"Sheet not found: {sheet_name}"}],
                        "warnings": [],
                    }
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Read data
            rows = []
            errors = []
            warnings = []

            skip_rows = options.get("skip_rows", 0)
            column_mapping = options.get("column_mapping", {})

            # Get headers from first row
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                headers.append(str(cell.value) if cell.value else f"Column{cell.column}")

            # Apply column mapping
            if column_mapping:
                mapped_headers = [column_mapping.get(h, h) for h in headers]
            else:
                mapped_headers = headers

            # Read data rows
            for i, row in enumerate(ws.iter_rows(min_row=2 + skip_rows)):
                try:
                    row_data = {}
                    for j, cell in enumerate(row):
                        if j < len(mapped_headers):
                            value = cell.value
                            row_data[mapped_headers[j]] = value

                    if any(v is not None for v in row_data.values()):
                        rows.append(row_data)

                except Exception as e:
                    errors.append({
                        "row": i + 2,
                        "column": "",
                        "message": str(e),
                    })

            wb.close()

            # Calculate quality score
            quality_score = 1.0 - (len(errors) / max(len(rows), 1))

            return {
                "success": len(errors) == 0,
                "rows_imported": len(rows),
                "data": rows,
                "headers": mapped_headers,
                "sheets": wb.sheetnames,
                "quality_score": quality_score,
                "errors": errors,
                "warnings": warnings,
            }

        except ImportError:
            return {
                "success": False,
                "error": "openpyxl library not installed",
                "rows_imported": 0,
                "errors": [{"row": 0, "column": "", "message": "openpyxl library required for Excel files"}],
                "warnings": [],
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "rows_imported": 0,
                "errors": [{"row": 0, "column": "", "message": str(e)}],
                "warnings": [],
            }

    def export_to_csv(self, data: List[Dict[str, Any]], file_path: str) -> Dict[str, Any]:
        """
        Export data to a CSV file.

        Args:
            data: List of dictionaries to export
            file_path: Output file path

        Returns:
            Dictionary with export results
        """
        try:
            if not data:
                return {
                    "success": False,
                    "error": "No data to export",
                    "rows_exported": 0,
                }

            # Get all unique headers
            headers = set()
            for row in data:
                headers.update(row.keys())
            headers = sorted(headers)

            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(data)

            return {
                "success": True,
                "rows_exported": len(data),
                "file_path": file_path,
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "rows_exported": 0,
            }

    def parse_station_data(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Parse imported data into station format.

        Args:
            data: Raw imported data

        Returns:
            List of station definitions
        """
        stations = []

        for row in data:
            station = {
                "id": str(row.get("id", row.get("station_id", f"station_{len(stations)}"))),
                "name": row.get("name", row.get("station_name", f"Station {len(stations) + 1}")),
            }

            # Parse cycle time
            cycle_time = row.get("cycle_time", row.get("cycleTime"))
            if cycle_time is not None:
                if isinstance(cycle_time, (int, float)):
                    station["cycle_time"] = {
                        "type": "constant",
                        "parameters": {"value": float(cycle_time)},
                    }
                elif isinstance(cycle_time, dict):
                    station["cycle_time"] = cycle_time

            # Parse optional parameters
            if "setup_time" in row or "setupTime" in row:
                setup_time = row.get("setup_time", row.get("setupTime"))
                if isinstance(setup_time, (int, float)):
                    station["setup_time"] = {
                        "type": "constant",
                        "parameters": {"value": float(setup_time)},
                    }

            if "mtbf" in row:
                station["mtbf"] = float(row["mtbf"])

            if "mttr" in row:
                station["mttr"] = float(row["mttr"])

            if "scrap_rate" in row or "scrapRate" in row:
                station["scrap_rate"] = float(row.get("scrap_rate", row.get("scrapRate", 0)))

            stations.append(station)

        return stations

    def detect_data_type(self, data: List[Dict[str, Any]]) -> str:
        """
        Detect the type of imported data.

        Args:
            data: Imported data

        Returns:
            Detected data type: "stations", "products", "schedule", "historical", or "unknown"
        """
        if not data:
            return "unknown"

        sample = data[0]
        keys = set(k.lower() for k in sample.keys())

        # Check for station data
        station_keys = {"cycle_time", "cycletime", "station", "station_id", "mtbf", "mttr"}
        if keys & station_keys:
            return "stations"

        # Check for product data
        product_keys = {"product", "product_id", "routing", "arrival_rate"}
        if keys & product_keys:
            return "products"

        # Check for schedule data
        schedule_keys = {"shift", "start_time", "end_time", "operator"}
        if keys & schedule_keys:
            return "schedule"

        # Check for historical data
        historical_keys = {"timestamp", "date", "time", "value", "measurement"}
        if keys & historical_keys:
            return "historical"

        return "unknown"
